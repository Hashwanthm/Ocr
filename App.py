from pathlib import Path
import re
import uuid

import pdfplumber
from flask import Flask, render_template, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# ======================================================
# CLEANING
# ======================================================

def clean_text(value):
    if value is None:
        return ""

    value = str(value)
    value = value.replace("\n", " ")
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def amount_to_number(value):
    value = clean_text(value)

    if value in ["", "-", ".", "0"]:
        return 0.0

    value = value.replace(",", "")
    value = re.sub(r"[^\d.]", "", value)

    if value == "":
        return 0.0

    try:
        return float(value)
    except Exception:
        return 0.0


def find_value(text, patterns, default=""):
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return clean_text(match.group(1))
    return default


# ======================================================
# PDF EXTRACTION
# ======================================================

def extract_pdf_text(pdf_path):
    all_text = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            all_text.append(text)

    return "\n".join(all_text)


def extract_pdf_tables(pdf_path):
    all_rows = []

    line_settings = {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "snap_tolerance": 3,
        "join_tolerance": 3,
        "intersection_tolerance": 5,
        "text_x_tolerance": 2,
        "text_y_tolerance": 3,
    }

    text_settings = {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_tolerance": 3,
        "join_tolerance": 3,
        "intersection_tolerance": 5,
        "text_x_tolerance": 2,
        "text_y_tolerance": 3,
    }

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables(table_settings=line_settings)

            if not tables:
                tables = page.extract_tables(table_settings=text_settings)

            for table in tables:
                for row in table:
                    cleaned_row = [clean_text(cell) for cell in row]

                    if any(cleaned_row):
                        all_rows.append(cleaned_row)

    return all_rows


# ======================================================
# PARSE BASIC DETAILS
# Works for:
# 1. Form 3A PDF
# 2. EPFO passbook PDF
# ======================================================

def parse_basic_details(text):
    normalized = clean_text(text)

    details = {
        "account_no": "",
        "uan": "",
        "name": "",
        "father_name": "",
        "factory_address": "",
        "establishment_id": "",
        "company_name": "",
        "statutory_rate": "12%",
        "higher_rate_employee": "NO",
        "higher_rate_employer": "NO",
        "higher_rate_pension": "NO",
        "date": "",
    }

    # Member ID / Name from EPFO passbook
    member_id = find_value(normalized, [
        r"Member\s*ID\s*/\s*Name\s*([A-Z0-9]+)\s*/",
        r"Member\s*ID\s*[:\-]?\s*([A-Z0-9]+)",
    ])

    member_name = find_value(normalized, [
        r"Member\s*ID\s*/\s*Name\s*[A-Z0-9]+\s*/\s*([A-Z][A-Z\s\.]+?)(?:\s*Date\s*of\s*Birth|\s*UAN|\s*EPF\s*Passbook|$)",
    ])

    # Establishment ID / Name from EPFO passbook
    establishment_id = find_value(normalized, [
        r"Establishment\s*ID\s*/\s*Name\s*([A-Z0-9]+)\s*/",
        r"Establishment\s*ID\s*([A-Z0-9]+)",
    ])

    establishment_name = find_value(normalized, [
        r"Establishment\s*ID\s*/\s*Name\s*[A-Z0-9]+\s*/\s*([A-Z0-9\s\.\-&PRIVATE LIMITED]+?)(?:\s*Member\s*ID|\s*Member|$)",
    ])

    # Form 3A account number
    account_no = find_value(normalized, [
        r"1\s*Account\s*No\.?\s*([A-Z0-9]+)",
        r"Account\s*No\.?\s*([A-Z0-9]+)",
    ])

    details["account_no"] = account_no or member_id
    details["establishment_id"] = establishment_id

    details["uan"] = find_value(normalized, [
        r"\bUAN\b\s*([0-9]{8,20})",
    ])

    # Form 3A name
    form_name = find_value(normalized, [
        r"2\s*Name\s*/\s*Surname\s*([A-Z][A-Z\s\.]+?)(?:\s*\(in\s*Block|\s*3\s*Father|\s*Father)",
        r"Name\s*/\s*Surname\s*([A-Z][A-Z\s\.]+?)(?:\s*\(in\s*Block|\s*3\s*Father|\s*Father)",
    ])

    details["name"] = form_name or member_name

    details["father_name"] = find_value(normalized, [
        r"3\s*Father'?s\s*/\s*Husband'?s\s*Name\s*([A-Z][A-Z\s\.]+?)(?:\s*4\s*Name|\s*Name\s*&\s*Address)",
        r"Father'?s\s*/\s*Husband'?s\s*Name\s*([A-Z][A-Z\s\.]+?)(?:\s*4\s*Name|\s*Name\s*&\s*Address)",
    ])

    factory_address = find_value(normalized, [
        r"4\s*Name\s*&\s*Address\s*of\s*the\s*Factory\s*(.*?)(?:Establishment\s*ID)",
        r"Name\s*&\s*Address\s*of\s*the\s*Factory\s*(.*?)(?:Establishment\s*ID)",
    ])

    details["factory_address"] = factory_address or establishment_name
    details["company_name"] = establishment_name

    if not details["company_name"] and factory_address:
        details["company_name"] = factory_address.split(",")[0].strip()

    if not details["company_name"]:
        details["company_name"] = "TRICORE SOLUTIONS PRIVATE LIMITED"

    statutory_rate = find_value(normalized, [
        r"Statutory\s*rate\s*of\s*Contribution\s*([0-9]+%?)",
    ])

    if statutory_rate:
        if "%" not in statutory_rate:
            statutory_rate += "%"
        details["statutory_rate"] = statutory_rate

    details["date"] = find_value(normalized, [
        r"Dated\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})",
        r"Printed\s*On\s*[:\-]?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})",
    ])

    return details


# ======================================================
# PARSE MONTHLY TABLE
# Works for Form 3A and EPFO passbook
# ======================================================

MONTH_PATTERN = r"^(Jan|January|Feb|February|Mar|March|Apr|April|May|Jun|June|Jul|July|Aug|August|Sep|September|Oct|October|Nov|November|Dec|December)"


def is_month_text(value):
    value = clean_text(value)
    return bool(re.match(MONTH_PATTERN, value, re.IGNORECASE))


def numeric_cells(cells):
    nums = []

    for cell in cells:
        text = clean_text(cell)

        if text == "":
            continue

        # Pick values like 15000, 15,000, 1800, 1,800, 0, 0.00
        if re.fullmatch(r"-?\d+(?:,\d{3})*(?:\.\d+)?", text):
            nums.append(amount_to_number(text))

    return nums


def parse_monthly_rows_from_tables(table_rows):
    monthly_rows = []

    for row in table_rows:
        if not row:
            continue

        cells = [clean_text(cell) for cell in row]

        if not cells:
            continue

        first_cell = cells[0]

        if not is_month_text(first_cell):
            continue

        row_text = " ".join(cells).upper()

        # EPFO passbook has CR / DR.
        # We use only CR contribution rows.
        if " DR " in f" {row_text} ":
            continue

        is_passbook_row = " CR " in f" {row_text} " or "CONTR" in row_text or "DUE-MONTH" in row_text

        if is_passbook_row:
            nums = numeric_cells(cells)

            # Expected passbook numeric sequence:
            # wage_epf, wage_eps, employee_contribution, employer_contribution, pension
            if len(nums) >= 5:
                monthly_rows.append({
                    "month": first_cell,
                    "wages": nums[0],
                    "epf": nums[2],
                    "epf_833": nums[3],
                    "pension": nums[4],
                    "refund": 0.0,
                    "non_contribution_days": 0.0,
                    "remarks": "",
                })
            elif len(nums) >= 3:
                monthly_rows.append({
                    "month": first_cell,
                    "wages": nums[0],
                    "epf": nums[1],
                    "epf_833": nums[2],
                    "pension": 0.0,
                    "refund": 0.0,
                    "non_contribution_days": 0.0,
                    "remarks": "",
                })

        else:
            # Form 3A table row:
            # Month, Wages, EPF, EPF 8 1/3, Pension, Refund, Days, Remarks
            while len(cells) < 8:
                cells.append("")

            monthly_rows.append({
                "month": cells[0],
                "wages": amount_to_number(cells[1]),
                "epf": amount_to_number(cells[2]),
                "epf_833": amount_to_number(cells[3]),
                "pension": amount_to_number(cells[4]),
                "refund": amount_to_number(cells[5]),
                "non_contribution_days": amount_to_number(cells[6]),
                "remarks": cells[7],
            })

    return monthly_rows


def parse_monthly_rows_from_text(text):
    monthly_rows = []

    for line in text.splitlines():
        line = clean_text(line)

        if not is_month_text(line):
            continue

        upper_line = line.upper()

        if " DR " in f" {upper_line} ":
            continue

        # This fallback works for many selectable PDFs.
        amounts = re.findall(r"\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+\.\d{2}", line)

        if len(amounts) >= 4:
            first_amount_pos = line.find(amounts[0])
            month_part = line[:first_amount_pos].strip()

            values = [amount_to_number(x) for x in amounts]

            if " CR " in f" {upper_line} " or "CONTR" in upper_line:
                wages = values[0]
                epf = values[1] if len(values) > 1 else 0
                employer = values[2] if len(values) > 2 else 0
                pension = values[3] if len(values) > 3 else 0
            else:
                wages = values[0]
                epf = values[1] if len(values) > 1 else 0
                employer = values[2] if len(values) > 2 else 0
                pension = values[3] if len(values) > 3 else 0

            monthly_rows.append({
                "month": month_part,
                "wages": wages,
                "epf": epf,
                "epf_833": employer,
                "pension": pension,
                "refund": 0.0,
                "non_contribution_days": 0.0,
                "remarks": "",
            })

    return monthly_rows


def parse_monthly_rows(table_rows, text):
    monthly_rows = parse_monthly_rows_from_tables(table_rows)

    if monthly_rows:
        return monthly_rows

    return parse_monthly_rows_from_text(text)


# ======================================================
# SAFE EXCEL HELPERS
# Fixes:
# 'MergedCell' object attribute 'value' is read-only
# ======================================================

def get_real_cell(ws, cell_ref):
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            min_col, min_row, max_col, max_row = merged_range.bounds
            return ws.cell(row=min_row, column=min_col)

    return ws[cell_ref]


def set_cell(ws, cell_ref, value, bold=False, size=10, align="left"):
    cell = get_real_cell(ws, cell_ref)
    cell.value = value
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=True
    )
    return cell


def make_border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_border(ws, start_row, start_col, end_row, end_col):
    border = make_border()

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical="center",
                wrap_text=True
            )


# ======================================================
# CREATE FORM 3A EXCEL OUTPUT
# ======================================================

def create_form3a_excel(details, monthly_rows, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "CORRECT"

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25

    column_widths = {
        "A": 24,
        "B": 16,
        "C": 14,
        "D": 14,
        "E": 16,
        "F": 14,
        "G": 22,
        "H": 22,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row_no in range(1, 70):
        ws.row_dimensions[row_no].height = 22

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[12].height = 34
    ws.row_dimensions[13].height = 34
    ws.row_dimensions[14].height = 34
    ws.row_dimensions[17].height = 30
    ws.row_dimensions[18].height = 48
    ws.row_dimensions[19].height = 22

    # ==================================================
    # HEADER
    # ==================================================

    ws.merge_cells("A1:H1")
    set_cell(
        ws,
        "A1",
        "Form No. 3A (Revised) [ See Paragraphs 35 & 42 of the Employees' Provident Funds Scheme, 1952 ]",
        bold=True,
        size=11,
        align="center"
    )

    ws.merge_cells("A2:H2")
    set_cell(
        ws,
        "A2",
        "[ See Paragraph 19 of the Employees' Pension Scheme, 1995 ]",
        size=10,
        align="center"
    )

    ws.merge_cells("A3:H3")
    set_cell(
        ws,
        "A3",
        "FOR UNEXEMPTED ESTABLISHMENTS ONLY",
        bold=True,
        size=11,
        align="center"
    )

    ws.merge_cells("A4:H4")
    set_cell(
        ws,
        "A4",
        "Contribution Card for the Currency Period from 1st April 2025 to 31st March 2026",
        bold=True,
        size=11,
        align="center"
    )

    # ==================================================
    # TOP DETAILS
    # ==================================================

    set_cell(ws, "A6", "1 Account No.")
    ws.merge_cells("B6:D6")
    set_cell(ws, "B6", details.get("account_no", ""))

    set_cell(ws, "A7", "UAN")
    ws.merge_cells("B7:D7")
    set_cell(ws, "B7", details.get("uan", ""))

    set_cell(ws, "A8", "2 Name / Surname\n(in Block Capitals)")
    ws.merge_cells("B8:D9")
    set_cell(ws, "B8", details.get("name", ""))

    set_cell(ws, "A10", "3 Father's/Husband's\nName")
    ws.merge_cells("B10:D11")
    set_cell(ws, "B10", details.get("father_name", ""))

    set_cell(ws, "A12", "4 Name & Address\nof the Factory")
    ws.merge_cells("B12:D14")
    set_cell(ws, "B12", details.get("factory_address", ""))

    set_cell(ws, "A15", "Establishment ID")
    ws.merge_cells("B15:D15")
    set_cell(ws, "B15", details.get("establishment_id", ""))

    ws.merge_cells("E6:G6")
    set_cell(ws, "E6", "5 Statutory rate of Contribution")
    set_cell(ws, "H6", details.get("statutory_rate", "12%"), bold=True, align="center")

    ws.merge_cells("E8:G9")
    set_cell(ws, "E8", "6 Voluntary higher rate of employee's\ncontribution if any")
    set_cell(ws, "H8", details.get("higher_rate_employee", "NO"), bold=True, align="center")

    ws.merge_cells("E10:G11")
    set_cell(ws, "E10", "7 Employer contribution on higher wages to\nEPF [Para 26(6)]")
    set_cell(ws, "H10", details.get("higher_rate_employer", "NO"), bold=True, align="center")

    ws.merge_cells("E12:G13")
    set_cell(ws, "E12", "8 Voluntary contribution to Pension Fund")
    set_cell(ws, "H12", details.get("higher_rate_pension", "NO"), bold=True, align="center")

    # ==================================================
    # TABLE HEADER
    # ==================================================

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws.merge_cells("A17:A19")
    set_cell(ws, "A17", "Months", bold=True, align="center")

    ws.merge_cells("B17:C17")
    set_cell(ws, "B17", "Employee's Share", bold=True, align="center")

    ws.merge_cells("D17:E17")
    set_cell(ws, "D17", "Employer's Share", bold=True, align="center")

    ws.merge_cells("F17:F18")
    set_cell(ws, "F17", "Refund of\nAdvance", bold=True, align="center")

    ws.merge_cells("G17:G18")
    set_cell(ws, "G17", "No. of days /\nperiod of non-\ncontributing\nservice if any", bold=True, align="center")

    ws.merge_cells("H17:H18")
    set_cell(ws, "H17", "Remarks", bold=True, align="center")

    set_cell(ws, "B18", "Amount of\nWages", bold=True, align="center")
    set_cell(ws, "C18", "EPF", bold=True, align="center")
    set_cell(ws, "D18", "EPF 8 1/3%\nif any", bold=True, align="center")
    set_cell(ws, "E18", "Pension Fund\ncontribution\n8 1/3%", bold=True, align="center")

    column_numbers = ["1", "2", "3", "4(a)", "4(b)", "5", "6", "7"]

    for col_index, number in enumerate(column_numbers, start=1):
        cell_ref = ws.cell(row=19, column=col_index).coordinate
        set_cell(ws, cell_ref, number, bold=True, align="center")

    for row in ws.iter_rows(min_row=17, max_row=19, min_col=1, max_col=8):
        for cell in row:
            cell.fill = header_fill

    # ==================================================
    # MONTHLY DATA
    # ==================================================

    data_start = 20
    minimum_month_rows = 12
    total_data_rows = max(minimum_month_rows, len(monthly_rows))

    for i in range(total_data_rows):
        excel_row = data_start + i

        if i < len(monthly_rows):
            item = monthly_rows[i]
        else:
            item = {
                "month": "",
                "wages": 0,
                "epf": 0,
                "epf_833": 0,
                "pension": 0,
                "refund": 0,
                "non_contribution_days": 0,
                "remarks": "",
            }

        set_cell(ws, f"A{excel_row}", item["month"])

        ws[f"B{excel_row}"] = item["wages"]
        ws[f"C{excel_row}"] = item["epf"]
        ws[f"D{excel_row}"] = item["epf_833"]
        ws[f"E{excel_row}"] = item["pension"]
        ws[f"F{excel_row}"] = item["refund"]
        ws[f"G{excel_row}"] = item["non_contribution_days"]
        set_cell(ws, f"H{excel_row}", item["remarks"])

        for col in range(2, 8):
            cell = ws.cell(row=excel_row, column=col)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # ==================================================
    # TOTAL ROW
    # ==================================================

    total_row = data_start + total_data_rows

    set_cell(ws, f"A{total_row}", "TOTAL", bold=True)

    for col in range(2, 8):
        col_letter = get_column_letter(col)
        cell = ws[f"{col_letter}{total_row}"]
        cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})"
        cell.font = Font(bold=True)
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # ==================================================
    # CERTIFICATION AREA
    # ==================================================

    cert_row = total_row + 2

    ws.merge_cells(start_row=cert_row, start_column=1, end_row=cert_row + 2, end_column=5)
    set_cell(
        ws,
        f"A{cert_row}",
        "Certified that the total amount of contribution both shares indicated in this card i.e.\n"
        "has already been remitted in full in EPF A/c No.1 and Pension Fund A/c No.10.",
        align="left"
    )

    ws.merge_cells(start_row=cert_row, start_column=6, end_row=cert_row, end_column=8)
    formula_cell = set_cell(
        ws,
        f"F{cert_row}",
        f"=C{total_row}+D{total_row}+E{total_row}",
        bold=True,
        align="center"
    )
    formula_cell.number_format = '"RS." 0.00'

    cert_row_2 = cert_row + 5

    ws.merge_cells(start_row=cert_row_2, start_column=1, end_row=cert_row_2 + 2, end_column=8)
    set_cell(
        ws,
        f"A{cert_row_2}",
        "Certified that the difference between the total of the contributions shown under columns 3 and 4(a) "
        "and 4(b) of the above table and that arrived at on the total wages shown in Column 2 at the prescribed "
        "rate is solely due to rounding off of contribution to the nearest rupee under the rules.",
        align="left"
    )

    sign_row = cert_row_2 + 5

    company_name = details.get("company_name") or "TRICORE SOLUTIONS PRIVATE LIMITED"

    ws.merge_cells(start_row=sign_row, start_column=5, end_row=sign_row, end_column=8)
    set_cell(
        ws,
        f"E{sign_row}",
        f"For {company_name},",
        bold=True,
        align="center"
    )

    set_cell(ws, f"A{sign_row + 3}", "Dated")
    set_cell(ws, f"B{sign_row + 3}", details.get("date", ""), bold=True)

    ws.merge_cells(start_row=sign_row + 3, start_column=6, end_row=sign_row + 3, end_column=8)
    set_cell(ws, f"F{sign_row + 3}", "Authorised Signatory", bold=True, align="center")

    note_row = sign_row + 5

    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 3, end_column=8)
    set_cell(
        ws,
        f"A{note_row}",
        "Note:- In respect of Form 3(A) sent to the Regional Office during the Course of the Currency period "
        "for the purpose of final settlement of the accounts of the member who had left service details of date "
        "and reason for leaving service should be furnished under column 7(a) & (b).\n"
        "In respect of those who are not members of the Pension Fund the employer's share of contribution to the EPF "
        "will be 8-1/3 of 10% as the case may be and is to be shown under Column 4(a).",
        align="left"
    )

    final_row = note_row + 3

    apply_border(ws, 1, 1, final_row, 8)

    ws.print_area = f"A1:H{final_row}"

    wb.save(excel_path)


# ======================================================
# FLASK ROUTES
# ======================================================

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/convert", methods=["POST"])
def convert_pdf():
    if "pdf" not in request.files:
        return jsonify({"error": "No PDF file received"}), 400

    file = request.files["pdf"]

    if file.filename == "":
        return jsonify({"error": "Please select a PDF file"}), 400

    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400

    safe_name = secure_filename(file.filename)
    unique_id = uuid.uuid4().hex

    uploaded_pdf_name = f"{unique_id}_{safe_name}"
    uploaded_pdf_path = UPLOAD_DIR / uploaded_pdf_name
    file.save(uploaded_pdf_path)

    try:
        text = extract_pdf_text(uploaded_pdf_path)
        table_rows = extract_pdf_tables(uploaded_pdf_path)

        details = parse_basic_details(text)
        monthly_rows = parse_monthly_rows(table_rows, text)

        if not monthly_rows:
            return jsonify({
                "error": "Monthly contribution table was not detected. If this PDF is scanned/image-based, OCR is required."
            }), 400

        excel_name = f"{Path(safe_name).stem}_form3a_{unique_id}.xlsx"
        excel_path = OUTPUT_DIR / excel_name

        create_form3a_excel(details, monthly_rows, excel_path)

        preview_rows = [
            ["Field", "Value"],
            ["Account No", details.get("account_no", "")],
            ["UAN", details.get("uan", "")],
            ["Name", details.get("name", "")],
            ["Father/Husband Name", details.get("father_name", "")],
            ["Establishment ID", details.get("establishment_id", "")],
            [""],
            ["Month", "Wages", "EPF", "EPF 8 1/3", "Pension", "Refund", "Days", "Remarks"],
        ]

        for item in monthly_rows:
            preview_rows.append([
                item["month"],
                item["wages"],
                item["epf"],
                item["epf_833"],
                item["pension"],
                item["refund"],
                item["non_contribution_days"],
                item["remarks"],
            ])

        return jsonify({
            "message": "Form 3A Excel created successfully",
            "rows": preview_rows,
            "download_url": f"/download/{excel_name}"
        })

    except Exception as error:
        return jsonify({"error": str(error)}), 500


@app.route("/download/<filename>")
def download_excel(filename):
    return send_from_directory(
        OUTPUT_DIR,
        filename,
        as_attachment=True
    )


if __name__ == "__main__":
    app.run(debug=True)
